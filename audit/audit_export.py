import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def export_audit_csv(get_conn, start_date, end_date):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    action,
                    actor,
                    target,
                    channel,
                    message_id,
                    detail,
                    created_at
                FROM audit_logs
                WHERE created_at::date BETWEEN %s AND %s
                ORDER BY created_at ASC
            """, (start_date, end_date))

            rows = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "action",
        "actor",
        "target",
        "channel",
        "message_id",
        "detail",
        "created_at"
    ])

    for row in rows:
        writer.writerow(row)

    output.seek(0)
    return output, len(rows)

def export_audit_xlsx(get_conn, start_date, end_date):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    action,
                    actor,
                    target,
                    channel,
                    message_id,
                    detail,
                    created_at
                FROM audit_logs
                WHERE created_at::date BETWEEN %s AND %s
                ORDER BY created_at ASC
            """, (start_date, end_date))
            rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = [
        "Action",
        "Actor",
        "Target",
        "Channel",
        "Message ID",
        "Detail",
        "Created At"
    ]

    # ===== Header style =====
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    center_align = Alignment(vertical="center")

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # ===== Data rows =====
    for row in rows:
        ws.append([
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6].strftime("%Y-%m-%d %H:%M:%S")
        ])

    # ===== Freeze header =====
    ws.freeze_panes = "A2"

    # ===== Auto width =====
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    filename = f"audit_{start_date}_{end_date}.xlsx"
    path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(path)

    return path, len(rows)